from django.http import HttpResponse
from nornir.core.task import Task, Result
from ipaddress import ip_network, ip_address
from django.contrib import messages
from .models import ScheduledTask
from .forms import ScheduledTaskForm
from .tasks import schedule_task
from .models import Device
from django.utils.timezone import now, timedelta
from django.utils import timezone
from django.db.models import Sum
from dateutil.relativedelta import relativedelta
import random
import plotly.graph_objects as go
from django.core import serializers
from django import forms
from django.forms import DateInput
from .permissions import IsAuthorOrReadOnly 
import plotly.graph_objects as go
import json
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .serializers import DeviceSerializer 
import asyncio
from puresnmp import Client, V2C, PyWrapper
from django.http import JsonResponse
from network_automation_app.models import Device, Log
import datetime
from datetime import datetime, timedelta
import netmiko
from pysnmp.hlapi import *
from pysnmp.smi import builder, view
from netmiko import ConnectHandler, SCPConn
import os, re, openpyxl, time, pickle
from datetime import datetime
from django.db.models import Q
from nornir import InitNornir
from django.conf import global_settings
from nornir_netmiko import netmiko_send_command, netmiko_send_config, netmiko_save_config
from openpyxl import load_workbook
from nornir.core.filter import F
from django.shortcuts import get_object_or_404
from openpyxl.styles import Border, Side, Alignment, PatternFill
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
import ipaddress
import logging
logging.basicConfig(level=logging.DEBUG)


def login_view(request):
    if request.method == 'POST':
        # 处理登录请求
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            return render(request, 'login.html', {'error_message': '用户名或密码错误'})
    else:
        # 渲染登录页面
        return render(request, 'login.html')

@login_required
def home(request):
    all_device = Device.objects.all()
    cisco_device = Device.objects.filter(Q(platform='cisco_ios') | Q(platform='cisco_ios_telnet'))
    FT_device = Device.objects.filter(platform='fortinet')
    NS_device = Device.objects.filter(platform='netscaler')
    F5_device = Device.objects.filter(platform='f5_tmsh')
    huawei_device = Device.objects.filter(platform='huawei')
    last_10_event = Log.objects.all().order_by('-id')[:10]
    context = {'all_device': len(all_device),
               'cisco_device': len(cisco_device),
               'huawei_device': len(huawei_device),
               'FT_device': len(FT_device),
               'NS_device': len(NS_device),
               'F5_device': len(F5_device),
               'last_10_event': last_10_event
               }
    return render(request, 'home.html', context)

def nornir_hosts(request):
    try:
        all_device = Device.objects.all()
        with open(global_settings.hostsyaml, 'w+', encoding='utf-8') as file:
            for device in all_device:
                file.write(str(device.hostname) + ':\n')
                file.write('    ' + 'hostname:' + ' ' + device.ip_address + '\n')
                file.write('    ' + 'platform:' + ' ' + device.platform + '\n')
                # 判断 platform 并添加相应的 group
                if device.platform.lower() == 'linux':
                    file.write("    groups:\n")
                    file.write("        - linux\n")
                else:
                    file.write("    groups:\n")
                    file.write("        - netdevice\n")
                file.write('    ' + 'data:' + '\n')
                file.write('        ' + 'level:' + ' ' + device.role + '\n')
                file.write('        ' + 'model:' + ' ' + device.model + '\n')
                file.write('\n')
        log = Log(target="nornir_host更新", action="nornir_host更新", status='Success', time=datetime.now(), messages='No Error')
        log.save()
    except Exception as e:
        log = Log(target="nornir_host更新", action="nornir_host更新", status='Error', time=datetime.now(), messages=e)
        log.save()
    with open(global_settings.hostsyaml, 'r') as file:
        result = file.read()
        return render(request, 'verify_config.html', {'result': result})

@login_required
def devices(request):
    all_devices = Device.objects.all().order_by('model')
    vendors = all_devices.values_list('model', flat=True).distinct()
    context = {
        'all_devices': all_devices,
        'vendors': vendors
    }
    return render(request, 'devices.html', context)


@login_required
def config(request):
    if request.method == 'GET':
        all_devices = Device.objects.all().order_by('model')
        vendors = all_devices.values_list('model', flat=True).distinct()
        context = {
            'all_devices': all_devices,
            'vendors': vendors
        }
        return render(request, 'config.html', context)
    elif request.method == 'POST':
        result = []
        selected_device_id = request.POST.getlist('device')
        command = request.POST['command'].splitlines()
        os.chdir(global_settings.inventory_path)
        for x in selected_device_id:
            try:
                dev = get_object_or_404(Device, pk=x)
                result.append(f'{dev.hostname}上的运行结果')
                nr = InitNornir(config_file="config.yaml")
                device = nr.filter(hostname=dev.ip_address)
                if dev.platform.lower() == 'huawei':
                    def CS(task):
                        commands = []
                        for cmd in command:
                            commands.append(cmd)
                        commands.append("commit")
                        task.run(netmiko_send_config, config_commands=commands)
                        task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")
                else:
                    def CS(task):
                        commands = []
                        for cmd in command:
                            commands.append(cmd)
                        task.run(netmiko_send_config, config_commands=commands)
                        task.run(netmiko_save_config)
                HW = device.run(task=CS)
                for sw in HW.keys():
                    for i in range(1, (len(command) + 1)):
                        result.append(HW[sw][i].result)
                log = Log(target=dev.ip_address, action='Configure', status='Success', time=datetime.now(), messages='No Error')
                log.save()
            except Exception as e:
                log = Log(target=dev.hostname, action='Configure', status='Error', time=datetime.now(), messages=e)
                log.save()
        result = '\n'.join(result)
        return render(request, 'verify_config.html', {'result': result})

@login_required
def verify_config(request):
    if request.method == 'GET':
        all_devices = Device.objects.all().order_by('model')
        vendors = all_devices.values_list('model', flat=True).distinct()
        context = {
            'all_devices': all_devices,
            'vendors': vendors
        }
        return render(request, 'config.html', context)

    elif request.method == 'POST':
        result = []
        selected_device_id = request.POST.getlist('device')
        command = request.POST['command'].splitlines()
        os.chdir(global_settings.inventory_path)
        for x in selected_device_id:
            try:
                dev = get_object_or_404(Device, pk=x)
                result.append(f'{dev.hostname}上的运行结果')
                nr = InitNornir(config_file="config.yaml")
                huawei = nr.filter(hostname=dev.ip_address)
                def HW(task):
                    for cmd in command:
                        task.run(netmiko_send_command, command_string=cmd)
                HW = huawei.run(task=HW)
                for sw in HW.keys():
                    for i in range(1, (len(command) + 1)):
                        result.append("%s,命令结果如下:\n" % command[(i - 1)])
                        result.append(HW[sw][i].result)
                        result.append("\n")
                log = Log(target=dev.ip_address, action='查看配置', status='Success', time=datetime.now(),
                          messages='No Error')
                log.save()
            except Exception as e:
                log = Log(target=dev.ip_address, action='查看配置', status='Error', time=datetime.now(), messages=e)
                log.save()
        result = '\n'.join(result)
        return render(request, 'verify_config.html', {'result': result})

@login_required
def backup_config(request):
    if request.method == 'GET':
        all_devices = Device.objects.all().order_by('model')
        vendors = all_devices.values_list('model', flat=True).distinct()
        context = {
            'all_devices': all_devices,
            'vendors': vendors
        }
        return render(request, 'backup_config.html', context)

    elif request.method == 'POST':
        result = []
        selected_device_id = request.POST.getlist('device')
        os.chdir(global_settings.inventory_path)
        for x in selected_device_id:
            try:
                dev = get_object_or_404(Device, pk=x)
                nr = InitNornir(config_file="config.yaml")
                device = nr.filter(hostname=dev.ip_address)
                netmiko_options = nr.inventory.defaults.connection_options.get("netmiko", {})
                extras = getattr(netmiko_options, "extras", {})
                secret = extras.get("secret", "")
                # 获取用户名、密码和秘密密钥
                username = nr.inventory.defaults.username
                password = nr.inventory.defaults.password
                connection_info = {
                    'device_type': 'f5_linux',
                    'ip': dev.ip_address,
                    'username': username,
                    'password': password,
                    'secret': secret,
                }
                if dev.platform.lower() == 'huawei':
                    def HW(task):
                        task.run(netmiko_send_command, command_string="dis curren")
                    HW = device.run(task=HW)
                    with open(f"{global_settings.network_config}{dev.hostname}.txt", "w") as f:
                        for sw in HW.keys():
                            f.write(HW[sw][1].result)
                if dev.platform.lower() == 'cisco_ios':
                    def CS(task):
                        task.run(netmiko_send_command, command_string="show run")
                    CS = device.run(task=CS)
                    with open(f"{global_settings.network_config}{dev.hostname}.txt", "w") as f:
                        for sw in CS.keys():
                            f.write(CS[sw][1].result)
                if dev.platform.lower() == 'cisco_ios_telnet':
                    def CS(task):
                        task.run(netmiko_send_command, command_string="show run")
                    CS = device.run(task=CS)
                    with open(f"{global_settings.network_config}{dev.hostname}.txt", "w") as f:
                        for sw in CS.keys():
                            f.write(CS[sw][1].result)
                if dev.platform.lower() == 'f5_tmsh':
                    def F5(task):
                        task.run(netmiko_send_command, command_string="show running-config")
                    F5 = device.run(task=F5)
                    with open(f"{global_settings.network_config}{dev.hostname}.txt", "w") as f:
                        for sw in F5.keys():
                            f.write(F5[sw][1].result)
                    try:
                        with ConnectHandler(**connection_info) as connect:
                            connect.send_command('rm -rf /var/local/ucs/F5-backupfile.ucs', read_timeout=120)
                            connect.send_command('tmsh save sys ucs F5-backupfile.ucs', read_timeout=120)
                        with ConnectHandler(**connection_info) as connect:
                            scp_conn = SCPConn(connect)
                            s_file = '/var/local/ucs/F5-backupfile.ucs'
                            d_file = f'{global_settings.network_config}F5-backupfile.ucs'
                            scp_conn.scp_get_file(s_file, d_file)
                            scp_conn.close()
                    except (OSError, netmiko.NetmikoTimeoutException):
                        result.append("Can not connect to Device " + connection_info['ip'])
                    except (EOFError, netmiko.NetMikoAuthenticationException):
                        result.append(connection_info['ip'] + " uername or passwd is wrong!")
                    except (ValueError, netmiko.NetMikoAuthenticationException):
                        result.append(connection_info['ip'] + " enable passwd wrong!")
                if dev.platform.lower() == 'netscaler':
                    def NS(task):
                        task.run(netmiko_send_command, command_string="show run")
                    NS = device.run(task=NS)
                    with open(f"{global_settings.network_config}{dev.hostname}.txt", "w") as f:
                        for sw in NS.keys():
                            f.write(NS[sw][1].result)
                    try:
                        with ConnectHandler(**connection_info) as connect:
                            connect.send_command('rm system backup backupfile.tgz', read_timeout=120)
                            connect.send_command('create system backup backupfile -level full', read_timeout=120)
                        with ConnectHandler(**connection_info) as connect:
                            scp_conn = SCPConn(connect)
                            s_file = '/var/ns_sys_backup/backupfile.tgz'
                            d_file = f'{global_settings.network_config}NS-backupfile.tgz'
                            scp_conn.scp_get_file(s_file, d_file)
                            scp_conn.close()
                    except (OSError, netmiko.NetmikoTimeoutException):
                        result.append("Can not connect to Device " + connection_info['ip'])
                    except (EOFError, netmiko.NetMikoAuthenticationException):
                        result.append(connection_info['ip'] + " uername or passwd is wrong!")
                    except (ValueError, netmiko.NetMikoAuthenticationException):
                        result.append(connection_info['ip'] + " enable passwd wrong!")
                if dev.platform.lower() == 'fortinet':
                    def FT(task):
                        task.run(netmiko_send_command, command_string="show full-configuration")
                    FT = device.run(task=FT)
                    with open(f"{global_settings.network_config}{dev.hostname}.txt", "w") as f:
                        for sw in FT.keys():
                            f.write(FT[sw][1].result)
                log = Log(target=dev.ip_address, action='Backup Configuration', status='Success', time=datetime.now(), messages='No Error')
                log.save()
            except Exception as e:
                log = Log(target=dev.ip_address, action='Backup Configuration', status='Error', time=datetime.now(), messages=e)
                log.save()
                result.append(f'{dev.hostname}配置备份失败，请查看日志!')
        directory = global_settings.network_config
        files = os.listdir(directory)
        num_files = len(files)
        for filename in files:
            filepath = os.path.join(directory, filename)
            filesize = os.path.getsize(filepath)
            file_modify_time = os.path.getmtime(filepath)
            # 转换日期格式
            file_modify_time = datetime.fromtimestamp(file_modify_time)
            # 添加文件信息到列表中
            result.append({
                'name': filename,
                'size': filesize,
                'modify_time': file_modify_time,
            })

        # 渲染模板
        return render(request, 'List_file.html', {'file_list': result, 'num_files': num_files})

def log(request):
    logs = Log.objects.all()
    context = {'logs': logs}
    return render(request, 'log.html', context)


def mac_arp(request):
    try:
        os.chdir(global_settings.inventory_path)
        nr = InitNornir(config_file="config.yaml")
        huawei_HJ = nr.filter(level='HJ', platform='huawei')
        huawei_JR = nr.filter(level='JR', platform='huawei')
        cisco_HJ = nr.filter(level='HJ', platform='cisco_ios_telnet')
        cisco_JR = nr.filter(level='JR', platform='cisco_ios')
        XY_k8s = nr.filter(level='HJ-JR', platform='huawei')
        output = huawei_HJ.run(netmiko_send_command, command_string='dis arp dynamic', read_timeout=120)
        output1 = huawei_JR.run(netmiko_send_command, command_string='dis mac-add dy', read_timeout=120)
        output2 = cisco_HJ.run(netmiko_send_command, command_string='show arp dynamic', read_timeout=120)
        output3 = cisco_JR.run(netmiko_send_command, command_string='show mac address-table dynamic', read_timeout=120)
        output4 = XY_k8s.run(netmiko_send_command, command_string='dis arp dynamic', read_timeout=120)
        output5 = XY_k8s.run(netmiko_send_command, command_string='dis mac-add dy', read_timeout=120)

        huawei_HJ_table = {}
        huawei_JR_table = {}
        cisco_HJ_table = {}
        cisco_JR_table = {}
        XY_k8s_arp_table = {}
        XY_k8s_mac_table = {}

        for a, b in [(output, huawei_HJ_table), (output1, huawei_JR_table), (output2, cisco_HJ_table),
                     (output3, cisco_JR_table), (output4, XY_k8s_arp_table), (output5, XY_k8s_mac_table)]:
            for sw in a.keys():
                b[sw] = a[sw].result

        for a, b in [(huawei_HJ_table, 'huawei_HJ_table'), (huawei_JR_table, 'huawei_JR_table'),
                     (cisco_HJ_table, 'cisco_HJ_table'), (cisco_JR_table, 'cisco_JR_table'),
                     (XY_k8s_arp_table, 'XY_k8s_arp_table'), (XY_k8s_mac_table, 'XY_k8s_mac_table')]:
            with open(
                    global_settings.arp_mac_data + b + '.pkl',
                    'wb') as f:
                pickle.dump(a, f)
        log = Log(target="mac_arpn表采集", action='mac_arpn表采集', status='Success', time=datetime.now(),
                  messages='No Error')
        log.save()
    except Exception as e:
        log = Log(target="mac_arpn表采集", action='mac_arpn表采集', status='Error', time=datetime.now(), messages=e)
        log.save()
    file_list = []
    directory = global_settings.arp_mac_data
    files = os.listdir(directory)
    num_files = len(files)
    for filename in files:
        filepath = os.path.join(directory, filename)
        filesize = os.path.getsize(filepath)
        file_modify_time = os.path.getmtime(filepath)
        # 转换日期格式
        file_modify_time = datetime.fromtimestamp(file_modify_time)
        # 添加文件信息到列表中
        file_list.append({
            'name': filename,
            'size': filesize,
            'modify_time': file_modify_time,
        })

    # 渲染模板
    return render(request, 'List_file.html', {'file_list': file_list, 'num_files': num_files})


@login_required
def information_collection(request):
    return render(request, 'information_collection.html')

@login_required
def mac_location(request):
    if request.method == 'GET':
        return render(request, 'mac_location.html')
    elif request.method == 'POST':
        try:
            os.chdir(global_settings.arp_mac_data)
            with open('huawei_HJ_table.pkl', 'rb') as f:
                output = pickle.load(f)
            with open('huawei_JR_table.pkl', 'rb') as f:
                output1 = pickle.load(f)
            with open('cisco_HJ_table.pkl', 'rb') as f:
                output2 = pickle.load(f)
            with open('cisco_JR_table.pkl', 'rb') as f:
                output3 = pickle.load(f)
            with open('XY_k8s_arp_table.pkl', 'rb') as f:
                output4 = pickle.load(f)
            with open('XY_k8s_mac_table.pkl', 'rb') as f:
                output5 = pickle.load(f)
            ip_addr = str(request.POST['mac_location'])
            ip_list = ip_addr.split(".")
            result = []
            for num in ip_list:
                if len(ip_list) == 4 and num.isdigit() and 0 <= int(num) <= 255:
                    # 网关是华为设备，接入也是华为设备
                    for sw in output.keys():
                        a = output[sw]
                        y = re.compile(ip_addr + r' {1,4}((?<= ).*?(?= ))')
                        c = y.search(a)
                        if c != None:
                            e = c.group(1)
                            for sw in output1.keys():
                                b = output1[sw]
                                f = re.compile(e + r'.*?((?<=#).*?(?= ))')
                                k = re.compile(e + r'.*?((?<=   )(.*?)(?= {1,10}dynamic))')
                                d = f.search(b)
                                p = k.search(b)
                                if d != None:
                                    if d.group(1).strip() not in ("Eth-Trunk1", "Eth-Trunk2", "Eth-Trunk100"):
                                        result.append({
                                            'sw': sw,
                                            'port': d.group(1).strip(),
                                        })
                                    else:
                                        pass
                                elif p != None:
                                    if "." not in p.group(1).strip():
                                        if p.group(1).strip() not in ("Eth-Trunk1", "Eth-Trunk2", "Eth-Trunk100"):
                                            result.append({
                                                'sw': sw,
                                                'port': p.group(1).strip(),
                                            })
                                        else:
                                            pass
                                    else:
                                        pass
                                else:
                                    pass
                        else:
                            pass

                    # 网关是华为设备，接入是思科设备
                    for sw in output.keys():
                        a = output[sw]
                        y = re.compile(ip_addr + r' {1,4}((?<= ).*?(?= ))')
                        c = y.search(a)
                        if c != None:
                            e = c.group(1)
                            g = e.replace('-', '.')
                            for sw in output3.keys():
                                b = output3[sw]
                                f = re.compile(g + r'.*(?<=DYNAMIC     )(.*)')
                                d = f.search(b)
                                if d != None:
                                    if d.group(1).strip() not in ("Po1", "Po2", "Po3"):
                                        result.append({
                                            'sw': sw,
                                            'port': d.group(1).strip(),
                                        })
                                else:
                                    pass
                        else:
                            pass

                    # 网关是思科设备，接入也是思科设备
                    for sw in output2.keys():
                        a = output2[sw]
                        y = re.compile(ip_addr + r' {1,20}\d{1,5}.*?((?<=   ).*?(?= ))')
                        c = y.search(a)
                        if c != None:
                            e = c.group(1)
                            for sw in output3.keys():
                                b = output3[sw]
                                f = re.compile(e + r'.*(?<=DYNAMIC     )(.*)')
                                d = f.search(b)
                                if d != None:
                                    if d.group(1).strip() not in ("Po1", "Po2", "Po3"):
                                        result.append({
                                            'sw': sw,
                                            'port': d.group(1).strip(),
                                        })
                                    else:
                                        pass
                        else:
                            pass

                    # 网关是思科设备，接入是华为设备
                    for sw in output2.keys():
                        a = output2[sw]
                        y = re.compile(ip_addr + r' {1,20}\d{1,5}.*?((?<=   ).*?(?= ))')
                        c = y.search(a)
                        if c != None:
                            e = c.group(1)
                            g = e.replace('.', '-')
                            for sw in output1.keys():
                                b = output1[sw]
                                f = re.compile(e + r'.*?((?<=#).*?(?= ))')
                                k = re.compile(e + r'.*?((?<=   )(.*?)(?= {1,10}dynamic))')
                                d = f.search(b)
                                p = k.search(b)
                                if d != None:
                                    if d.group(1).strip() not in ("Eth-Trunk1", "Eth-Trunk2", "Eth-Trunk100"):
                                        result.append({
                                            'sw': sw,
                                            'port': d.group(1).strip(),
                                        })
                                    else:
                                        pass
                                elif p != None:
                                    if p.group(1).strip() not in ("Eth-Trunk1", "Eth-Trunk2", "Eth-Trunk100"):
                                        result.append({
                                            'sw': sw,
                                            'port': p.group(1).strip(),
                                        })
                                    else:
                                        pass
                                else:
                                    pass
                        else:
                            pass

                    # XY_k8s集群:
                    for sw in output4.keys():
                        a = output4[sw]
                        y = re.compile(ip_addr + r' {1,4}((?<= ).*?(?= ))')
                        c = y.search(a)
                        if c != None:
                            e = c.group(1)
                            for sw in output5.keys():
                                b = output5[sw]
                                f = re.compile(e + r'.*?((?<=#).*?(?= ))')
                                k = re.compile(e + r'.*?((?<=   )(.*?)(?= {1,10}dynamic))')
                                d = f.search(b)
                                p = k.search(b)

                                if d != None:
                                    if d.group(1).strip() not in ("Eth-Trunk1", "Eth-Trunk2", "Eth-Trunk100"):
                                        result.append({
                                            'sw': sw,
                                            'port': d.group(1).strip(),
                                        })
                                    else:
                                        pass
                                elif p != None:
                                    if "." not in p.group(1).strip():
                                        if p.group(1).strip() not in ("Eth-Trunk1", "Eth-Trunk2", "Eth-Trunk100"):
                                            result.append({
                                                'sw': sw,
                                                'port': p.group(1).strip(),
                                            })
                                        else:
                                            pass
                                    else:
                                        pass
                                else:
                                    pass
                        else:
                            pass
                else:
                    result.append({
                        'sw': "地址不合法,请确认ip地址输入是否正确",
                        'port': None,
                    })
            log = Log(target="dev.ip_address", action='服务器接入位置查询', status='Success', time=datetime.now(), messages='No Error')
            log.save()
        except Exception as e:
            log = Log(target="dev.ip_address", action='服务器接入位置查询', status='Error', time=datetime.now(), messages=e)
            log.save()
    # 将字典转换为元组
    list_of_tuples = [tuple(d.items()) for d in result]
    # 去重
    unique_list_of_tuples = list(set(list_of_tuples))
    # 将元组转换为字典
    result = [dict(t) for t in unique_list_of_tuples]
    return render(request, 'verify_config_mac_location.html', {'result': result})




@login_required
def inspection(request):
    try:
        os.chdir(global_settings.device_information)
        # 边框
        thin = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

        # 纵向靠左  水平居中
        align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 纵向居中  水平居中
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        os.chdir(f"{global_settings.device_information}")

        if os.path.exists("./设备巡检/设备巡检表.xlsx") == False:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "HW8860"
            wb.create_sheet(title="CE12800")
            wb.create_sheet(title="CE68|58")
            wb.create_sheet(title="NE20")
            wb.create_sheet(title="HW5700")
            wb.create_sheet(title="CS4500")
            wb.create_sheet(title="CS")
            wb.save("./设备巡检/设备巡检表.xlsx")
        else:
            os.remove("./设备巡检/设备巡检表.xlsx")
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "HW8860"
            wb.create_sheet(title="CE12800")
            wb.create_sheet(title="CE68|58")
            wb.create_sheet(title="NE20")
            wb.create_sheet(title="HW5700")
            wb.create_sheet(title="CS4500")
            wb.create_sheet(title="CS")
            wb.save("./设备巡检/设备巡检表.xlsx")

        os.chdir(global_settings.inventory_path)
        nr = InitNornir(config_file="config.yaml")
        HW88 = nr.filter(F(platform='huawei') and F(model='HW8800'))
        HW128 = nr.filter(F(platform='huawei') and F(model='HW12800'))
        NE20 = nr.filter(F(platform='huawei') and F(model='HWNE20'))
        HW68 = nr.filter(F(model='HW5800') | F(model='HW6800'))
        HW57 = nr.filter(model='HW5700')
        CS45 = nr.filter(model='CS45')
        CS = nr.filter(model='CS')

        HW88_card = re.compile(r'\d\/\d +(\S+) +', re.M)
        HW12800_card = re.compile(r'-      (CE-L\S+)', re.M)
        NE20_card = re.compile(r'\d+\/\d+ +\S+ +(\w+) +', re.M)
        CS45_card = re.compile(r'(WS-\w+-\w+\SE)', re.M)

        cmds = ['dis ver', 'dis esn', 'dis device card', 'dis memory', 'dis cpu', 'dis device fan', 'dis device power',
                'dis device alarm hardware']
        HW12800_cmds = ['dis ver', 'dis device board', 'dis memory', 'dis cpu', 'dis device fan', 'dis device power',
                        'dis device alarm hardware', 'dis esn']
        NE20_cmds = ['dis ver', 'dis esn', 'dis device pic-status', 'dis memory-usage', 'dis cpu-usage', 'dis fan',
                     'dis power', 'dis alarm hardware']
        HW68_cmds = ['dis ver', 'dis memory', 'dis cpu', 'dis device fan', 'dis device power',
                     'dis device alarm hardware',
                     'dis esn']
        HW57_cmds = ['dis ver', 'dis memory', 'dis cpu', 'dis device', 'dis alarm active',
                     'dis device manufacture-info',
                     'dis power']
        CS45_cmds = ['show version', 'show inventory', 'show processes cpu', 'show processes memory', 'show module',
                     'show environment alarm', 'show environment status']
        CS_cmds = ['show version', 'show environment all', 'show processes cpu', 'show processes memory']

        def HW8800(task):
            for cmd in cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def HW12800(task):
            for cmd in HW12800_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def HWNE20(task):
            for cmd in NE20_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def HW6800(task):
            for cmd in HW68_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def HW5700(task):
            for cmd in HW57_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def CS4500(task):
            for cmd in CS45_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def CSJR(task):
            for cmd in CS_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        HW88 = HW88.run(task=HW8800)
        HW128 = HW128.run(task=HW12800)
        HWNE20 = NE20.run(task=HWNE20)
        HW68 = HW68.run(task=HW6800)
        HW57 = HW57.run(task=HW5700)
        CS45 = CS45.run(task=CS4500)
        CS = CS.run(task=CSJR)

        def HW_CE_XJ(sw_num, x, e, sheet):
            model = re.compile(r'(?<=HUAWEI )(\S{5,20})(?= )|cisco (\S+) \(')
            CS45_model = re.compile(r'cisco (\S+) \(')
            SN = re.compile(r'(?:ESN \S+ \S+ \d:|ESN \S+ \S+ \d :|ESN \S+ \S+:|\d +- {1,4}) (\w{6,20})', re.M)
            SN_CS45 = re.compile(r'Cisco Systems,.*\n.*SN: (\S+)', re.M)
            SN_CS = re.compile(r'System Serial Number +: (\S+)', re.M)
            uptime = re.compile(r'((?<=uptime is ).*)')
            cpu = re.compile(
                r'(?:System cpu use rate is :|System CPU Using Percentage :|CPU Usage            :|CPU utilization for five seconds:) +(\d+\S)')
            memory = re.compile(r'Memory Using Percentage(?: Is:|:) (\d+\S)')
            memory_CS45 = re.compile(r'(?<=System memory  : )(.*)(?=, \S+ kernel reserved)')
            memory_CS = re.compile(r'Processor Pool Total: +(\d+ Used: +\d+ Free: +\d+)')
            fan = re.compile(r'FAN\d +\S+ +(\S+)', re.M)
            fan_CS45 = re.compile(r'Fantray : ([a-zA-Z]+)')
            fan_CS = re.compile(r'FAN \S+ is (\w+)')
            Power = re.compile(r'PWR\d +(\S+) +AC', re.M)
            Power_CS45 = re.compile(r'PS\d +\S+ +AC +\S+ +(\w+)', re.M)
            Power_CS = re.compile(r'\d\w  (\S{3,15}) +', re.M)
            fan = list(set(fan.findall(txt)))
            power = list(set(Power.findall(txt)))
            fan_CS45 = list(set(fan_CS45.findall(txt)))
            power_CS45 = list(set(Power_CS45.findall(txt)))
            fan_CS = list(set(fan_CS.findall(txt)))
            power_CS = list(set(Power_CS.findall(txt)))
            if not fan and not power:
                if power_CS45 and fan_CS45:
                    fan = "\n".join(map(str, fan_CS45))
                    power = "\n".join(map(str, power_CS45))
                else:
                    if power_CS and fan_CS:
                        fan = "\n".join(map(str, fan_CS))
                        power = "\n".join(map(str, power_CS))
                    else:
                        fan = "——"
                        power = "——"
            else:
                fan = "\n".join(map(str, fan))
                power = "\n".join(map(str, power))
            alarm = re.compile(r'month', re.M)
            if CS45_model.search(txt):
                model = CS45_model.search(txt)
            else:
                model = model.search(txt)
            if SN.findall(txt):
                SN = SN.findall(txt)
            else:
                if SN_CS45.findall(txt):
                    SN = SN_CS45.findall(txt)
                else:
                    SN = SN_CS.findall(txt)
            uptime = uptime.search(txt)
            cpu = cpu.search(txt)
            if memory.search(txt):
                memory = memory.search(txt)
            else:
                if memory_CS.search(txt):
                    memory = memory_CS.search(txt)
                else:
                    memory = memory_CS45.search(txt)
            alarm = alarm.findall(txt)
            SN = "\n".join(map(str, SN))
            if not alarm:
                alarm = "本月无硬件告警"
            else:
                pass
            if x < (sw_num + 1):
                sheet.merge_cells(start_row=e, end_row=e, start_column=1, end_column=4)
                sheet.cell(row=e, column=1).value = hostname
                sheet.cell(row=e, column=1).alignment = center
                for l in range(e, e + 6):
                    for i in range(1, 5):
                        sheet.cell(row=l, column=i).border = thin
                for i in range(e + 1, e + 6):
                    for l in range(1, 5):
                        sheet.cell(row=i, column=l).alignment = align
                sheet.cell(row=e + 1, column=1).value = "设备型号"
                sheet.cell(row=e + 1, column=2).value = model.group(1)
                sheet.cell(row=e + 1, column=3).value = "运行时间"
                sheet.cell(row=e + 1, column=4).value = uptime.group(1)
                sheet.cell(row=e + 2, column=1).value = "序列号"
                sheet.cell(row=e + 2, column=2).value = SN
                sheet.cell(row=e + 2, column=3).value = "板卡信息"
                sheet.cell(row=e + 2, column=4).value = card
                sheet.cell(row=e + 3, column=1).value = "CPU_usage"
                sheet.cell(row=e + 3, column=2).value = cpu.group(1)
                sheet.cell(row=e + 3, column=3).value = "Memory_usage"
                sheet.cell(row=e + 3, column=4).value = memory.group(1)
                sheet.cell(row=e + 4, column=1).value = "电源状态"
                sheet.cell(row=e + 4, column=2).value = power
                sheet.cell(row=e + 4, column=3).value = "风扇状态"
                sheet.cell(row=e + 4, column=4).value = fan
                sheet.merge_cells(start_row=e + 5, end_row=e + 5, start_column=1, end_column=4)
                sheet.cell(row=e + 5, column=1).value = alarm
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 25
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 28
            wb.save("./设备巡检/设备巡检表.xlsx")

        os.chdir(global_settings.device_information)
        wb = load_workbook("./设备巡检/设备巡检表.xlsx")
        for a in (HW88, HW128, HWNE20, HW68, HW57, CS45, CS):
            e = 1
            x = 1
            for sw in a.keys():
                hostname = sw
                txt = a[sw][(1)].result
                if a == HW88:
                    for i in range(1, len(cmds)):
                        txt = txt + a[sw][(i + 1)].result
                    sheet = wb['HW8860']
                    card = HW88_card.findall(txt)
                    Demo_dict = {}
                    for item in card:
                        Demo_dict.update({item: card.count(item)})
                    c = re.sub("'", "", str(Demo_dict).strip('{}'))
                    card = "\n".join(map(str, c.split(', ')))
                    HW_CE_XJ(len(a.keys()), x, e, sheet)
                    e += 8
                    x += 1
                elif a == HW128:
                    for i in range(1, len(HW12800_cmds)):
                        txt = txt + a[sw][(i + 1)].result
                    sheet = wb['CE12800']
                    card = HW12800_card.findall(txt)
                    Demo_dict = {}
                    for item in card:
                        Demo_dict.update({item: card.count(item)})
                    c = re.sub("'", "", str(Demo_dict).strip('{}'))
                    card = "\n".join(map(str, c.split(', ')))
                    HW_CE_XJ(len(a.keys()), x, e, sheet)
                    e += 8
                    x += 1
                elif a == HWNE20:
                    for i in range(1, len(NE20_cmds)):
                        txt = txt + a[sw][(i + 1)].result
                    sheet = wb['NE20']
                    card = NE20_card.findall(txt)
                    Demo_dict = {}
                    for item in card:
                        Demo_dict.update({item: card.count(item)})
                    c = re.sub("'", "", str(Demo_dict).strip('{}'))
                    card = "\n".join(map(str, c.split(', ')))
                    HW_CE_XJ(len(a.keys()), x, e, sheet)
                    e += 8
                    x += 1
                elif a == HW68:
                    for i in range(1, len(HW68_cmds)):
                        txt = txt + a[sw][(i + 1)].result
                    sheet = wb['CE68|58']
                    card = "None"
                    HW_CE_XJ(len(a.keys()), x, e, sheet)
                    e += 8
                    x += 1
                elif a == HW57:
                    for i in range(1, len(HW57_cmds)):
                        txt = txt + a[sw][(i + 1)].result
                    sheet = wb['HW5700']
                    card = "None"
                    HW_CE_XJ(len(a.keys()), x, e, sheet)
                    e += 8
                    x += 1
                elif a == CS45:
                    for i in range(1, len(CS45_cmds)):
                        txt = txt + a[sw][(i + 1)].result
                    sheet = wb['CS4500']
                    card = CS45_card.findall(txt)
                    Demo_dict = {}
                    for item in card:
                        Demo_dict.update({item: card.count(item)})
                    c = re.sub("'", "", str(Demo_dict).strip('{}'))
                    card = "\n".join(map(str, c.split(', ')))
                    HW_CE_XJ(len(a.keys()), x, e, sheet)
                    e += 8
                    x += 1
                elif a == CS:
                    for i in range(1, len(CS_cmds)):
                        txt = txt + a[sw][(i + 1)].result
                    sheet = wb['CS']
                    card = "None"
                    HW_CE_XJ(len(a.keys()), x, e, sheet)
                    e += 8
                    x += 1
                else:
                    pass
        log = Log(target="设备巡检", action='设备巡检', status='Success', time=datetime.now(), messages='No Error')
        log.save()
    except Exception as e:
        log = Log(target="设备巡检", action='设备巡检', status='Error', time=datetime.now(), messages=e)
        log.save()

    file_list = []
    directory = f'{global_settings.device_information}/设备巡检/'
    files = os.listdir(directory)
    num_files = len(files)
    for filename in files:
        filepath = os.path.join(directory, filename)
        filesize = os.path.getsize(filepath)
        file_modify_time = os.path.getmtime(filepath)
        # 转换日期格式
        file_modify_time = datetime.fromtimestamp(file_modify_time)
        # 添加文件信息到列表中
        file_list.append({
            'name': filename,
            'size': filesize,
            'modify_time': file_modify_time,
        })

    # 渲染模板
    return render(request, 'List_file.html', {'file_list': file_list, 'num_files': num_files})


def network_version(request):
    try:
        os.chdir(f"{global_settings.device_information}")

        if os.path.exists("./设备巡检/设备版本号.xlsx") == False:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "NET_Version"
            wb.save("./设备巡检/设备版本号.xlsx")
        else:
            os.remove("./设备巡检/设备版本号.xlsx")
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "NET_Version"
            wb.save("./设备巡检/设备版本号.xlsx")

        os.chdir(f"{global_settings.inventory_path}")
        nr = InitNornir(config_file="config.yaml")
        HW = nr.filter(F(platform='huawei'))
        CS = nr.filter(F(platform='cisco_ios'))
        CS_telnet = nr.filter(F(platform='cisco_ios_telnet'))
        F5 = nr.filter(F(platform='f5_tmsh'))
        FT = nr.filter(F(platform='fortinet'))
        NS = nr.filter(F(platform='netscaler'))

        F5_cmds = ['show sys hardware']
        NS_cmds = ['show version', 'show hardware']
        FT_cmds = ['get system status']
        HW_cmds = ['dis version']
        CS_cmds = ['show version']
        CS_telnet_cmds = ['show version']

        def HW1(task):
            for cmd in HW_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def CS1(task):
            for cmd in CS_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def CS1_telnet(task):
            for cmd in CS_telnet_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def F51(task):
            for cmd in F5_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def FT1(task):
            for cmd in FT_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        def NS1(task):
            for cmd in NS_cmds:
                task.run(netmiko_send_command, command_string=cmd)

        HW = HW.run(task=HW1)
        CS = CS.run(task=CS1)
        CS_telnet = CS_telnet.run(task=CS1_telnet)
        F5 = F5.run(task=F51)
        FT = FT.run(task=FT1)
        NS = NS.run(task=NS1)

        Hostname = []
        Ip = []
        IOS_Version = []
        Platform = []
        for sw in HW.keys():
            output = HW[sw][(1)].result
            for i in range(1, len(HW_cmds)):
                output = output + HW[sw][(i + 1)].result
            a = re.search(r'(?<=Version ).*(?= \()', output)
            b = re.search(r'(?<=HUAWEI )\S{5,}(?= )', output)
            ip = Device.objects.filter(hostname=sw).values_list('ip_address', flat=True).first()
            Ip.append(ip)
            Hostname.append(sw)
            if b:
                Platform.append(b.group())
            else:
                Platform.append("None")
            if a:
                IOS_Version.append(a.group())
            else:
                IOS_Version.append("None")

        for sw in CS.keys():
            output = CS[sw][(1)].result
            for i in range(1, len(CS_cmds)):
                output = output + CS[sw][(i + 1)].result
            b = re.search(r'(?<=cisco ).*?(?= \()', output)
            a = re.search(r'(?<=Version ).*(?= RELEASE)', output)
            ip = Device.objects.filter(hostname=sw).values_list('ip_address', flat=True).first()
            Ip.append(ip)
            Hostname.append(sw)
            if b:
                Platform.append(b.group())
            else:
                Platform.append("None")
            if a:
                IOS_Version.append(a.group())
            else:
                IOS_Version.append("None")

        for sw in CS_telnet.keys():
            output = CS_telnet[sw][(1)].result
            for i in range(1, len(CS_telnet_cmds)):
                output = output + CS_telnet[sw][(i + 1)].result
            a = re.search(r'(?<=Version ).*(?= RELEASE)', output)
            b = re.search(r'(?<=cisco ).*?(?= \()', output)
            ip = Device.objects.filter(hostname=sw).values_list('ip_address', flat=True).first()
            Ip.append(ip)
            Hostname.append(sw)
            if b:
                Platform.append(b.group())
            else:
                Platform.append("None")
            if a:
                IOS_Version.append(a.group())
            else:
                IOS_Version.append("None")

        for sw in FT.keys():
            output = FT[sw][(1)].result
            for i in range(1, len(FT_cmds)):
                output = output + FT[sw][(i + 1)].result
            a = re.search(r'(?<=D ).*(?= \()', output)
            b = re.search(r'(?<=Version: ).*?(?= )', output)
            ip = Device.objects.filter(hostname=sw).values_list('ip_address', flat=True).first()
            Ip.append(ip)
            Hostname.append(sw)
            if b:
                Platform.append(b.group())
            else:
                Platform.append("None")
            if a:
                IOS_Version.append(a.group())
            else:
                IOS_Version.append("None")

        for sw in NS.keys():
            output = NS[sw][(1)].result
            for i in range(1, len(NS_cmds)):
                output = output + NS[sw][(i + 1)].result
            a = re.search(r'(?<=NetScaler ).*(?=, Date)', output)
            b = re.search(r'(?<=Platform: ).*?(?= )', output)
            ip = Device.objects.filter(hostname=sw).values_list('ip_address', flat=True).first()
            Ip.append(ip)
            Hostname.append(sw)
            if b:
                Platform.append(b.group())
            else:
                Platform.append("None")
            if a:
                IOS_Version.append(a.group())
            else:
                IOS_Version.append("None")

        for sw in F5.keys():
            output = F5[sw][(1)].result
            for i in range(1, len(F5_cmds)):
                output = output + F5[sw][(i + 1)].result
            a = re.search(r'(?<=Build: ).*(?= )', output)
            b = re.search(r'(?<=Name           ).*', output)
            ip = Device.objects.filter(hostname=sw).values_list('ip_address', flat=True).first()
            Ip.append(ip)
            Hostname.append(sw)
            if b:
                Platform.append(b.group())
            else:
                Platform.append("None")
            if a:
                IOS_Version.append(a.group())
            else:
                IOS_Version.append("None")

        os.chdir(f"{global_settings.device_information}")
        wb = openpyxl.load_workbook('./设备巡检/设备版本号.xlsx')
        sheet = wb['NET_Version']
        sheet.delete_rows(2, 200)
        sheet.append(['Hostname', 'Ip', 'Platform', 'IOS_Version'])
        for i in range(len(Hostname)):
            sheet.append([Hostname[i], Ip[i], Platform[i], IOS_Version[i]])
        yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        sheet['A1'].fill = yellowFill
        sheet['B1'].fill = yellowFill
        sheet['C1'].fill = yellowFill
        sheet['D1'].fill = yellowFill
        for row in sheet.rows:
            for cell in row:
                # 设置单元格边框
                cell.border = thin_border
                # 如果单元格有值，更新列宽
                if cell.value:
                    sheet.column_dimensions[cell.column_letter].auto_size = True

        wb.save('./设备巡检/设备版本号.xlsx')
        log = Log(target="设备版本型号采集", action='设备版本型号采集', status='Success', time=datetime.now(), messages='No Error')
        log.save()
    except Exception as e:
        log = Log(target="设备版本型号采集", action='设备版本型号采集', status='Error', time=datetime.now(), messages=e)
        log.save()

    file_list = []
    directory = f'{global_settings.device_information}/设备巡检/'
    files = os.listdir(directory)
    num_files = len(files)
    for filename in files:
        filepath = os.path.join(directory, filename)
        filesize = os.path.getsize(filepath)
        file_modify_time = os.path.getmtime(filepath)
        # 转换日期格式
        file_modify_time = datetime.fromtimestamp(file_modify_time)
        # 添加文件信息到列表中
        file_list.append({
            'name': filename,
            'size': filesize,
            'modify_time': file_modify_time,
        })

    return render(request, 'List_file.html', {'file_list': file_list, 'num_files': num_files})

@login_required
def FD_SFTP_white(request):
    if request.method == 'GET':
        return render(request, 'FD_SFTP_white.html')
    elif request.method == 'POST':
        FD_SFTP_white1 = str(request.POST['FD_SFTP_white1'])
        FD_SFTP_white2 = str(request.POST['FD_SFTP_white2'])
        FD_SFTP_white3 = str(request.POST['FD_SFTP_white3'])
        FD_SFTP_white4 = str(request.POST['FD_SFTP_white4'])
        Match = []
        Edit = []
        result = []
        if not FD_SFTP_white1 and not FD_SFTP_white2:
            if not FD_SFTP_white3:
                if FD_SFTP_white4:
                    device = {
                        'device_type': 'linux',
                        'host': '172.16.226.28',
                        'username': 'root',
                        'password': '51nbops868#',
                    }
                    try:
                        net_connect = ConnectHandler(**device)
                        output = net_connect.send_command('sh /root/create-account.sh {}'.format(FD_SFTP_white4))  # 发送命令示例
                        net_connect.disconnect()
                        result = output.split('\n')
                        result.insert(0, '添加账号为:{}'.format(FD_SFTP_white4))
                        result = '\n'.join(result)
                    except Exception as e:
                        print(f"连接失败: {e}")
                else:
                    result.append("输入为空，请核实后再次输入")
            else:
                os.chdir(f"{global_settings.inventory_path}")
                nr = InitNornir(config_file="config.yaml")
                FT = nr.filter(F(platform='fortinet'))
                commands = ['config vdom', 'edit IDC_temp', 'config firewall address', 'show']
                def List(task):
                    for cmd in commands:
                        task.run(netmiko_send_config, config_commands=cmd)
                FT = FT.run(task=List)
                for sw in FT.keys():
                    for i in range(1, (len(commands) + 1)):
                        Match.append(FT[sw][i].result)
                        Match.append("\n")
                Match = '\n'.join(Match)
                a = re.search(fr'edit "(.*)"(\n.*set.comment.*?)?\n.*set subnet {FD_SFTP_white3}', Match)
                if f"subnet {FD_SFTP_white3}" in Match:
                    a = a.group(1)
                    os.chdir(f"{global_settings.inventory_path}")
                    nr = InitNornir(config_file="config.yaml")
                    FT = nr.filter(F(platform='fortinet'))
                    Edit = ['config vdom', 'edit IDC_temp', 'config firewall addrgrp', 'edit sftp', 'unselect member {}'.format(a), 'end']
                    def List(task):
                        for cmd in Edit:
                            task.run(netmiko_send_config, config_commands=cmd)
                    FT = FT.run(task=List)
                    for sw in FT.keys():
                        for i in range(1, (len(Edit) + 1)):
                            result.append(FT[sw][i].result)
                            result.append("\n")
                    result = '\n'.join(result)
                else:
                    result.append("白名单地址不存在或已删除")

        else:
            try:
                os.chdir(f"{global_settings.inventory_path}")
                nr = InitNornir(config_file="config.yaml")
                FT = nr.filter(F(platform='fortinet'))
                commands = ['config vdom', 'edit IDC_temp', 'config firewall address', 'show']
                def List(task):
                    for cmd in commands:
                        task.run(netmiko_send_config, config_commands=cmd)
                FT = FT.run(task=List)
                for sw in FT.keys():
                    for i in range(1, (len(commands) + 1)):
                        Match.append(FT[sw][i].result)
                        Match.append("\n")
                Match = '\n'.join(Match)
                if FD_SFTP_white1 and FD_SFTP_white2:
                    ip2 = ipaddress.IPv4Address(FD_SFTP_white2)
                    subnet_mask2 = str(ipaddress.IPv4Interface(ip2).netmask)
                    if f"subnet {FD_SFTP_white1}" in Match:
                        a = re.search(fr'edit "(.*)"(\n.*set.comment.*?)?\n.*set subnet {FD_SFTP_white1}', Match)
                        a = a.group(1)
                        os.chdir(f"{global_settings.inventory_path}")
                        nr = InitNornir(config_file="config.yaml")
                        FT = nr.filter(F(platform='fortinet'))
                        print(a)
                        type(a)
                        Edit = ['config vdom', 'edit IDC_temp', 'config firewall address', 'edit {}'.format(a), 'set subnet {} {}'.format(FD_SFTP_white2, subnet_mask2), 'end']
                        def List(task):
                            for cmd in Edit:
                                task.run(netmiko_send_config, config_commands=cmd)
                        FT = FT.run(task=List)
                        for sw in FT.keys():
                            for i in range(1, (len(Edit) + 1)):
                                result.append(FT[sw][i].result)
                                result.append("\n")
                        result = '\n'.join(result)
                    else:
                        result.append("原白名单地址不存在,请核实")
                elif not FD_SFTP_white1 and FD_SFTP_white2:
                    ip2 = ipaddress.IPv4Address(FD_SFTP_white2)
                    subnet_mask2 = str(ipaddress.IPv4Interface(ip2).netmask)
                    if f"subnet {FD_SFTP_white2}" in Match:
                        result.append("白名单地址已经存在")
                    else:
                        if f"edit \"{FD_SFTP_white2}\"" in Match:
                            result.append("地址名称重复, 请手工添加")
                        else:
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT = nr.filter(F(platform='fortinet'))
                            Edit = ['config vdom', 'edit IDC_temp', 'config firewall address', 'edit {}'.format(FD_SFTP_white2), 'set subnet {} {}'.format(FD_SFTP_white2, subnet_mask2), 'end', 'config firewall addrgrp', 'edit sftp', 'append member {}'.format(FD_SFTP_white2), 'end']
                            def List(task):
                                for cmd in Edit:
                                    task.run(netmiko_send_config, config_commands=cmd)
                            FT = FT.run(task=List)
                            for sw in FT.keys():
                                for i in range(1, (len(Edit) + 1)):
                                    result.append(FT[sw][i].result)
                                    result.append("\n")
                            result = '\n'.join(result)
                elif FD_SFTP_white1 and not FD_SFTP_white2:
                    result.append("请输入要被替换的白名单地址")
                else:
                    result.append("Error")
                log = Log(target="FD_SFTP白名单修改", action='FD_SFTP白名单修改', status='Success', time=datetime.now(), messages='No Error')
                log.save()
            except Exception as e:
                log = Log(target="FD_SFTP白名单修改", action='FD_SFTP白名单修改', status='Error', time=datetime.now(), messages=e)
                log.save()
        return render(request, 'verify_config.html', {'result': result})

@login_required
def K8S_BGP(request):
    if request.method == 'GET':
        return render(request, 'K8S_BGP.html')
    elif request.method == 'POST':
        result = []
        Match = []
        delete_bgp_ip = str(request.POST['delete_bgp'])
        add_bgp_ip = str(request.POST['add_bgp'])
        DSJ_delete_bgp_ip = str(request.POST['DSJ_delete_bgp'])
        DSJ_add_bgp_ip = str(request.POST['DSJ_add_bgp'])
        XX_stable_k8s_delete_bgp = str(request.POST['XX_stable_k8s_delete_bgp'])
        XX_stable_k8s_add_bgp = str(request.POST['XX_stable_k8s_add_bgp'])
        D03D04_network = ['172.16.240.64/27', '172.16.240.96/27']
        C05C06_network = ['172.16.240.0/27', '172.16.240.32/27']
        DSJ_network = ['172.16.28.0/22']
        networks_all = ['172.16.240.64/27', '172.16.240.96/27', '172.16.240.0/27', '172.16.240.32/27']
        if not delete_bgp_ip and not add_bgp_ip and not DSJ_delete_bgp_ip and not DSJ_add_bgp_ip and not XX_stable_k8s_delete_bgp and not XX_stable_k8s_add_bgp:
            result.append("输入地址为空，请核实后再次输入")
        else:
            try:
                if XX_stable_k8s_delete_bgp:
                    try:
                        delete_address = ipaddress.ip_address(XX_stable_k8s_delete_bgp)
                        os.chdir(f"{global_settings.inventory_path}")
                        nr = InitNornir(config_file="config.yaml")
                        FT = nr.filter(F(hostname='172.17.0.84'))
                        commands = ['dis bgp peer']
                        def List(task):
                            task.run(netmiko_send_config, config_commands=commands)

                        FT = FT.run(task=List)
                        for sw in FT.keys():
                            for i in range(1, (len(commands) + 1)):
                                Match.append(FT[sw][i].result)
                                Match.append("\n")
                        Match = '\n'.join(Match)
                        a = re.search(fr' +{XX_stable_k8s_delete_bgp} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+', Match)
                        if a.group(1) == "Established":
                            result.append("请先node删除bgp再在交换机侧删除bgp")
                        else:
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            username = nr.inventory.defaults.username
                            password = nr.inventory.defaults.password
                            connection_info = {
                                'device_type': 'huawei',
                                'ip': '172.17.0.84',
                                'username': username,
                                'password': password,
                            }
                            with ConnectHandler(**connection_info) as connect:
                                connect.send_command(command_string='system-view', expect_string=r']')
                                connect.send_command(command_string='bgp 65534', expect_string=r']')
                                connect.send_command(command_string=f'undo peer {XX_stable_k8s_delete_bgp}',
                                                     expect_string=r']|:')
                                connect.send_command(command_string='Y\n', expect_string=r']')
                                connect.send_command(command_string='commit', expect_string=r']')
                            with ConnectHandler(**connection_info) as connect:
                                output = connect.send_command(command_string=f'dis bgp peer | in {XX_stable_k8s_delete_bgp}')
                                a = re.search(fr' +{XX_stable_k8s_delete_bgp} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+',
                                              output)
                                if a:
                                    result.append(output)
                                    result.append("未完成删除bgp邻居")
                                else:
                                    result.append("已完成删除,验证如下: ")
                                    result.append(f'dis bgp peer | in {XX_stable_k8s_delete_bgp}')
                                    result.append(output)
                    except ValueError:
                        result.append("地址不合法,请重新输入")   
                else:
                    pass
                 
                if XX_stable_k8s_add_bgp:
                    try:
                        add_address = ipaddress.ip_address(XX_stable_k8s_add_bgp)     
                        network = ipaddress.ip_network('10.248.1.0/24') 
                        network2 = ipaddress.ip_network('10.248.2.0/24')                            
                        if add_address in network:
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT = nr.filter(F(hostname='172.17.0.84'))
                            commands = ['dis bgp peer']

                            def List(task):
                                for cmd in commands:
                                    task.run(netmiko_send_config, config_commands=cmd)

                            FT = FT.run(task=List)
                            for sw in FT.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{XX_stable_k8s_add_bgp} +\d +\d+ +\d+ +\d+ +\d+ +\S+ ([a-zA-Z]+) +\d+', Match)
                            if a:
                                if a.group(1) == "Established":
                                    result.append(f"{XX_stable_k8s_add_bgp}bgp邻居已存在且状态为Established")
                                else:
                                    result.append(f"{XX_stable_k8s_add_bgp}bgp配置已存在")
                            else:
                                os.chdir(f"{global_settings.inventory_path}")
                                nr = InitNornir(config_file="config.yaml")
                                FT = nr.filter(F(hostname='172.17.0.84'))
                                commands = ['bgp 65534', f'peer {XX_stable_k8s_add_bgp} as-number 65538',  
                                            f'peer {XX_stable_k8s_add_bgp} connect-interface Vlanif24',  
                                            f'peer {XX_stable_k8s_add_bgp} group K8S',  
                                            'ipv4-family unicast',  
                                            f'peer {XX_stable_k8s_add_bgp} group K8S', 
                                            f'peer {XX_stable_k8s_add_bgp} enable',
                                            f'peer {XX_stable_k8s_add_bgp} route-policy Deny-all export', 'commit']            

                                def List(task):
                                    task.run(netmiko_send_config, config_commands=commands)
                                    task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")

                                HW = FT.run(task=List)
                                for sw in HW.keys():
                                    for i in range(1, (len(commands) + 1)):
                                        result.append(HW[sw][i].result)
                        elif add_address in network2:
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT = nr.filter(F(hostname='172.17.0.84'))
                            commands = ['dis bgp peer']

                            def List(task):
                                task.run(netmiko_send_config, config_commands=commands)

                            FT = FT.run(task=List)
                            for sw in FT.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{XX_stable_k8s_add_bgp} +\d +\d+ +\d+ +\d+ +\d+ +\S+ ([a-zA-Z]+) +\d+', Match)
                            if a:
                                if a.group(1) == "Established":
                                    result.append(f"{XX_stable_k8s_add_bgp}bgp邻居已存在且状态为Established")
                                else:
                                    result.append(f"{XX_stable_k8s_add_bgp}bgp配置已存在")
                            else:
                                os.chdir(f"{global_settings.inventory_path}")
                                nr = InitNornir(config_file="config.yaml")
                                FT = nr.filter(F(hostname='172.17.0.84'))
                                commands = ['bgp 65534', f'peer {XX_stable_k8s_add_bgp} as-number 65538',            
                                            f'peer {XX_stable_k8s_add_bgp} connect-interface Vlanif100',              
                                            f'peer {XX_stable_k8s_add_bgp} group K8S',                               
                                            'ipv4-family unicast',                                                   
                                            f'peer {XX_stable_k8s_add_bgp} group K8S',                               
                                            f'peer {XX_stable_k8s_add_bgp} enable',                                  
                                            f'peer {XX_stable_k8s_add_bgp} route-policy Deny-all export', 'commit']  

                                def List(task):
                                    a = task.run(netmiko_send_config, config_commands=commands,
                                                 expect_string="Continue?")
                                    a.connection.send_command_timing("Y")
                                    task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")

                                HW = FT.run(task=List)
                                for sw in HW.keys():
                                    for i in range(1, (len(commands) + 1)):
                                        result.append(HW[sw][i].result)
                        elif any(add_address not in ipaddress.ip_network(network) for network in networks_all):
                            result.append("地址网段有误，请检查地址网段是否存在")
                        else:
                            pass
                    except ValueError:
                        result.append("地址不合法,请重新输入")
                else:
                    pass
                if delete_bgp_ip:
                    try:
                        delete_address = ipaddress.ip_address(delete_bgp_ip)
                        if any(delete_address in ipaddress.ip_network(network) for network in D03D04_network):
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT = nr.filter(F(hostname='51.51.51.22'))
                            commands = ['dis bgp peer']

                            def List(task):
                                task.run(netmiko_send_config, config_commands=commands)

                            FT = FT.run(task=List)
                            for sw in FT.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+', Match)
                            if a.group(1) == "Established":
                                result.append("请先node删除bgp再在交换机侧删除bgp")
                            else:
                                os.chdir(f"{global_settings.inventory_path}")
                                nr = InitNornir(config_file="config.yaml")
                                username = nr.inventory.defaults.username
                                password = nr.inventory.defaults.password
                                connection_info = {
                                    'device_type': 'huawei',
                                    'ip': '51.51.51.22',
                                    'username': username,
                                    'password': password,
                                }
                                with ConnectHandler(**connection_info) as connect:
                                    connect.send_command(command_string='system-view', expect_string=r']')
                                    connect.send_command(command_string='bgp 65535', expect_string=r']')
                                    connect.send_command(command_string=f'undo peer {delete_bgp_ip}',
                                                         expect_string=r']|:')
                                    connect.send_command(command_string='Y\n', expect_string=r']')
                                    connect.send_command(command_string='commit', expect_string=r']')
                                with ConnectHandler(**connection_info) as connect:
                                    output = connect.send_command(command_string=f'dis bgp peer | in {delete_bgp_ip}')
                                    a = re.search(fr' +{delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+',
                                                  output)
                                    if a:
                                        result.append(output)
                                        result.append("未完成删除bgp邻居")
                                    else:
                                        result.append("已完成删除,验证如下: ")
                                        result.append(f'dis bgp peer | in {delete_bgp_ip}')
                                        result.append(output)
                        elif any(delete_address in ipaddress.ip_network(network) for network in C05C06_network):
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT = nr.filter(F(hostname='51.51.51.33'))
                            commands = ['dis bgp peer']

                            def List(task):
                                for cmd in commands:
                                    task.run(netmiko_send_config, config_commands=cmd)

                            FT = FT.run(task=List)
                            for sw in FT.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ ([a-zA-Z]+) +\d+', Match)
                            if a.group(1) == "Established":
                                result.append("请先node删除bgp再在交换机侧删除bgp")
                            else:
                                os.chdir(f"{global_settings.inventory_path}")
                                nr = InitNornir(config_file="config.yaml")
                                username = nr.inventory.defaults.username
                                password = nr.inventory.defaults.password
                                connection_info = {
                                    'device_type': 'huawei',
                                    'ip': '51.51.51.33',
                                    'username': username,
                                    'password': password,
                                }
                                with ConnectHandler(**connection_info) as connect:
                                    connect.send_command(command_string='system-view', expect_string=r']')
                                    connect.send_command(command_string='bgp 65535', expect_string=r']')
                                    connect.send_command(command_string=f'undo peer {delete_bgp_ip}',
                                                         expect_string=r']|:')
                                    connect.send_command(command_string='Y\n', expect_string=r']')
                                    connect.send_command(command_string='commit', expect_string=r']')
                                with ConnectHandler(**connection_info) as connect:
                                    output = connect.send_command(command_string=f'dis bgp peer | in {delete_bgp_ip}')
                                    a = re.search(fr' +{delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+',
                                                  output)
                                    if a:
                                        result.append(output)
                                        result.append("未完成删除bgp邻居")
                                    else:
                                        result.append("已完成删除,验证如下: ")
                                        result.append(f'dis bgp peer | in {delete_bgp_ip}')
                                        result.append(output)
                        elif any(delete_address not in ipaddress.ip_network(network) for network in networks_all):
                            result.append("地址网段有误，请检查地址网段是否存在")
                        else:
                            pass
                    except ValueError:
                        result.append("地址不合法,请重新输入")
                else:
                    pass
                if add_bgp_ip:
                    try:
                        add_address = ipaddress.ip_address(add_bgp_ip)
                        if any(add_address in ipaddress.ip_network(network) for network in D03D04_network):
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT = nr.filter(F(hostname='51.51.51.22'))
                            commands = ['dis bgp peer']

                            def List(task):
                                for cmd in commands:
                                    task.run(netmiko_send_config, config_commands=cmd)

                            FT = FT.run(task=List)
                            for sw in FT.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{add_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ ([a-zA-Z]+) +\d+', Match)
                            if a:
                                if a.group(1) == "Established":
                                    result.append(f"{add_bgp_ip}bgp邻居已存在且状态为Established")
                                else:
                                    result.append(f"{add_bgp_ip}bgp配置已存在")
                            else:
                                os.chdir(f"{global_settings.inventory_path}")
                                nr = InitNornir(config_file="config.yaml")
                                FT = nr.filter(F(hostname='51.51.51.22'))
                                commands = ['bgp 65535', f'peer {add_bgp_ip} as-number 65535', 'ipv4-family unicast',
                                            f'peer {add_bgp_ip} enable',
                                            f'peer {add_bgp_ip} route-policy xy-k8s-route-filter import',
                                            f'peer {add_bgp_ip} route-policy k8s-route-export-filter-all export',
                                            f'peer {add_bgp_ip} reflect-client', 'commit']

                                def List(task):
                                    task.run(netmiko_send_config, config_commands=commands)
                                    task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")

                                HW = FT.run(task=List)
                                for sw in HW.keys():
                                    for i in range(1, (len(commands) + 1)):
                                        result.append(HW[sw][i].result)
                        elif any(add_address in ipaddress.ip_network(network) for network in C05C06_network):
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT = nr.filter(F(hostname='51.51.51.33'))
                            commands = ['dis bgp peer']

                            def List(task):
                                task.run(netmiko_send_config, config_commands=commands)

                            FT = FT.run(task=List)
                            for sw in FT.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{add_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ ([a-zA-Z]+) +\d+', Match)
                            if a:
                                if a.group(1) == "Established":
                                    result.append(f"{add_bgp_ip}bgp邻居已存在且状态为Established")
                                else:
                                    result.append(f"{add_bgp_ip}bgp配置已存在")
                            else:
                                os.chdir(f"{global_settings.inventory_path}")
                                nr = InitNornir(config_file="config.yaml")
                                FT = nr.filter(F(hostname='51.51.51.33'))
                                commands = ['bgp 65535', f'peer {add_bgp_ip} as-number 65535', 'ipv4-family unicast',
                                            f'peer {add_bgp_ip} enable',
                                            f'peer {add_bgp_ip} route-policy xy-k8s-route-filter import',
                                            f'peer {add_bgp_ip} route-policy k8s-route-export-filter-all export',
                                            f'peer {add_bgp_ip} reflect-client', 'commit']

                                def List(task):
                                    a = task.run(netmiko_send_config, config_commands=commands,
                                                 expect_string="Continue?")
                                    a.connection.send_command_timing("Y")
                                    task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")

                                HW = FT.run(task=List)
                                for sw in HW.keys():
                                    for i in range(1, (len(commands) + 1)):
                                        result.append(HW[sw][i].result)
                        elif any(add_address not in ipaddress.ip_network(network) for network in networks_all):
                            result.append("地址网段有误，请检查地址网段是否存在")
                        else:
                            pass
                    except ValueError:
                        result.append("地址不合法,请重新输入")
                else:
                    pass
                if DSJ_delete_bgp_ip:
                    try:
                        DSJ_delete_address = ipaddress.ip_address(DSJ_delete_bgp_ip)
                        if any(DSJ_delete_address in ipaddress.ip_network(network) for network in DSJ_network):
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT1 = nr.filter(F(hostname='51.51.51.100'))
                            FT2 = nr.filter(F(hostname='51.51.51.101'))
                            commands = ['dis bgp peer']

                            def List(task):
                                task.run(netmiko_send_config, config_commands=commands)

                            FT1 = FT1.run(task=List)
                            FT2 = FT2.run(task=List)
                            Match = []
                            for sw in FT1.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT1[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{DSJ_delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+', Match)
                            if a:
                                if a.group(1) == "Established":
                                    result.append(f"FD4-D22-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp邻居已存在且状态为Established")
                                else:
                                    username = nr.inventory.defaults.username
                                    password = nr.inventory.defaults.password
                                    connection_info = {
                                        'device_type': 'huawei',
                                        'ip': '51.51.51.100',
                                        'username': username,
                                        'password': password,
                                    }
                                    with ConnectHandler(**connection_info) as connect:
                                        connect.send_command(command_string='system-view', expect_string=r']')
                                        connect.send_command(command_string='bgp 65533', expect_string=r']')
                                        connect.send_command(command_string=f'undo peer {DSJ_delete_bgp_ip}',
                                                             expect_string=r']|:')
                                        connect.send_command(command_string='Y\n', expect_string=r']')
                                        connect.send_command(command_string='commit', expect_string=r']')
                                    with ConnectHandler(**connection_info) as connect:
                                        output = connect.send_command(
                                            command_string=f'dis bgp peer | in {DSJ_delete_bgp_ip}')
                                        a = re.search(
                                            fr' +{DSJ_delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+',
                                            output)
                                        if a:
                                            result.append(output)
                                            result.append("FD4-D24-CE12812-DSJ-1未完成删除bgp邻居")
                                        else:
                                            result.append('FD4-D24-CE12812-DSJ-1已完成删除,验证如下: ')
                                            result.append(f'dis bgp peer | in {DSJ_delete_bgp_ip}')
                                            result.append(output)
                            else:
                                result.append("FD4-D22-CE12812-DSJ-1配置不存在")
                            Match = []
                            for sw in FT2.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match.append(FT2[sw][i].result)
                                    Match.append("\n")
                            Match = '\n'.join(Match)
                            a = re.search(fr' +{DSJ_delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+', Match)
                            if a:
                                if a.group(1) == "Established":
                                    result.append(f"FD4-D24-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp邻居已存在且状态为Established")
                                else:
                                    username = nr.inventory.defaults.username
                                    password = nr.inventory.defaults.password
                                    connection_info = {
                                        'device_type': 'huawei',
                                        'ip': '51.51.51.101',
                                        'username': username,
                                        'password': password,
                                    }
                                    with ConnectHandler(**connection_info) as connect:
                                        connect.send_command(command_string='system-view', expect_string=r']')
                                        connect.send_command(command_string='bgp 65533', expect_string=r']')
                                        connect.send_command(command_string=f'undo peer {DSJ_delete_bgp_ip}',
                                                             expect_string=r']|:')
                                        connect.send_command(command_string='Y\n', expect_string=r']')
                                        connect.send_command(command_string='commit', expect_string=r']')
                                    with ConnectHandler(**connection_info) as connect:
                                        output = connect.send_command(
                                            command_string=f'dis bgp peer | in {DSJ_delete_bgp_ip}')
                                        a = re.search(
                                            fr' +{DSJ_delete_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+',
                                            output)
                                        if a:
                                            result.append(output)
                                            result.append("FD4-D24-CE12812-DSJ-1未完成删除bgp邻居")
                                        else:
                                            result.append('FD4-D24-CE12812-DSJ-1已完成删除,验证如下: ')
                                            result.append(f'dis bgp peer | in {DSJ_delete_bgp_ip}')
                                            result.append(output)
                            else:
                                result.append("FD4-D24-CE12812-DSJ-1配置不存在")

                        elif any(DSJ_delete_address not in ipaddress.ip_network(network) for network in DSJ_network):
                            result.append("地址网段有误，请检查地址网段是否存在")
                        else:
                            pass
                    except ValueError:
                        result.append("地址不合法,请重新输入")
                else:
                    pass
                if DSJ_add_bgp_ip:
                    try:
                        DSJ_add_address = ipaddress.ip_address(DSJ_add_bgp_ip)
                        if any(DSJ_add_address in ipaddress.ip_network(network) for network in DSJ_network):
                            os.chdir(f"{global_settings.inventory_path}")
                            nr = InitNornir(config_file="config.yaml")
                            FT_D22 = nr.filter(F(hostname='51.51.51.100'))
                            FT_D24 = nr.filter(F(hostname='51.51.51.101'))
                            commands = ['dis bgp peer']
                            def List(task):
                                task.run(netmiko_send_config, config_commands=commands)
                            FT_D22_12812 = FT_D22.run(task=List)
                            FT_D24_12812 = FT_D24.run(task=List)
                            Match_D22 = []
                            Match_D24 = []
                            for sw in FT_D22_12812.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match_D22.append(FT_D22_12812[sw][i].result)
                                    Match_D22.append("\n")
                            Match_D22 = '\n'.join(Match_D22)
                            for sw in FT_D24_12812.keys():
                                for i in range(1, (len(commands) + 1)):
                                    Match_D24.append(FT_D24_12812[sw][i].result)
                                    Match_D24.append("\n")
                            Match_D24 = '\n'.join(Match_D24)
                            a_D22 = re.search(fr' +{DSJ_add_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+', Match_D22)
                            a_D24 = re.search(fr' +{DSJ_add_bgp_ip} +\d +\d+ +\d+ +\d+ +\d+ +\S+ +(\S+) +\d+', Match_D24)
                            if a_D22 and a_D24:
                                print("111_D22")
                                if a_D22.group(1) == "Established":
                                    result.append(f"FD4-D22-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp邻居已存在且状态为Established")
                                else:
                                    result.append(f"FD4-D22-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp配置已存在")
                                if a_D24.group(1) == "Established":
                                    result.append(f"FD4-D24-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp邻居已存在且状态为Established")
                                else:
                                    result.append(f"FD4-D24-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp配置已存在")
                            elif a_D22 == None and a_D24:
                                result.append(f"FD4-D24-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp配置已存在")
                                commands = ['bgp 65533', f'peer {DSJ_add_bgp_ip} as-number 64001',
                                            f'peer {DSJ_add_bgp_ip} connect-interface LoopBack0', 'ipv4-family unicast',
                                            f'peer {DSJ_add_bgp_ip} enable',
                                            f'peer {DSJ_add_bgp_ip} route-policy fd-k8s-route-filter import',
                                            f'peer {DSJ_add_bgp_ip} route-policy k8s-route-export-filter-all export',
                                            'commit']
                                def List_D22(task):
                                    task.run(netmiko_send_config, config_commands=commands)
                                    task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")
                                D22_12812 = FT_D22.run(task=List_D22)
                                for sw in D22_12812.keys():
                                    for i in range(1, (len(commands) + 1)):
                                        result.append(D22_12812[sw][i].result)
                            elif a_D24 == None and a_D22:
                                result.append(f"FD4-D22-CE12812-DSJ-1中{DSJ_add_bgp_ip}bgp配置已存在")
                                commands = ['bgp 65533', f'peer {DSJ_add_bgp_ip} as-number 64001',
                                            f'peer {DSJ_add_bgp_ip} connect-interface LoopBack0', 'ipv4-family unicast',
                                            f'peer {DSJ_add_bgp_ip} enable',
                                            f'peer {DSJ_add_bgp_ip} route-policy fd-k8s-route-filter import',
                                            f'peer {DSJ_add_bgp_ip} route-policy k8s-route-export-filter-all export',
                                            'commit']

                                def List_D24(task):
                                    task.run(netmiko_send_config, config_commands=commands)
                                    task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")

                                D24_12812 = FT_D24.run(task=List_D24)
                                for sw in D24_12812.keys():
                                    for i in range(1, (len(commands) + 1)):
                                        result.append(D24_12812[sw][i].result)
                            elif a_D22 == None and a_D24 == None:
                                print("1111111")
                                commands = ['bgp 65533', f'peer {DSJ_add_bgp_ip} as-number 64001',
                                            f'peer {DSJ_add_bgp_ip} connect-interface LoopBack0', 'ipv4-family unicast',
                                            f'peer {DSJ_add_bgp_ip} enable',
                                            f'peer {DSJ_add_bgp_ip} route-policy fd-k8s-route-filter import',
                                            f'peer {DSJ_add_bgp_ip} route-policy k8s-route-export-filter-all export',
                                            'commit']
                                os.chdir(f"{global_settings.inventory_path}")
                                nr = InitNornir(config_file="config.yaml")
                                FT = nr.filter(F(hostname='51.51.51.100') | F(hostname='51.51.51.101'))
                                def List(task):
                                    task.run(netmiko_send_config, config_commands=commands)
                                    task.run(netmiko_save_config, cmd="save", confirm="True", confirm_response="y")
                                FD12812 = FT.run(task=List)
                                for sw in FD12812.keys():
                                    for i in range(1, (len(commands) + 1)):
                                        result.append(FD12812[sw][i].result)
                        elif any(DSJ_add_address not in ipaddress.ip_network(network) for network in DSJ_network):
                            result.append("地址网段有误，请检查地址网段是否存在")
                        else:
                            pass
                    except ValueError:
                        result.append("地址不合法,请重新输入")
                else:
                    pass
                log = Log(target="K8S—BGP操作", action='K8S—BGP操作', status='Success', time=datetime.now(), messages='No Error')
                log.save()
            except Exception as e:
                log = Log(target="K8S—BGP操作", action='K8S—BGP操作', status='Error', time=datetime.now(), messages=e)
                log.save()
        result = '\n'.join(result)
        return render(request, 'verify_config.html', {'result': result})

async def get_device_info(request):
   client = PyWrapper(Client('172.16.37.33', V2C("51zhangdan")))
   output = await client.get("1.3.6.1.2.1.1.1.0")
   return output

class List(generics.ListCreateAPIView): 
    queryset = Device.objects.all() 
    serializer_class = DeviceSerializer

    
class Detail(generics.RetrieveUpdateDestroyAPIView): 
    permission_classes = (IsAuthorOrReadOnly,) 
    queryset = Device.objects.all() 
    serializer_class = DeviceSerializer


class Current(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        username = request.user.username
        return Response({'username': username}) 

@login_required
def XY_internet(request):
    return render(request, 'XY_internet.html')



def task_list(request):
    tasks = ScheduledTask.objects.all()
    return render(request, 'task_list.html', {'tasks': tasks})

def add_task(request):
    script_choices = [
        (f, f) for f in os.listdir('/root/django_network_automation/network_automation_app') if f.endswith('.py')
    ]

    if request.method == 'POST':
        form = ScheduledTaskForm(request.POST)
        form.fields['script'].choices = script_choices
        if form.is_valid():
            script = form.cleaned_data['script']
            cron_schedule = form.cleaned_data['cron_schedule']
            ScheduledTask.objects.create(script_path=script, cron_schedule=cron_schedule, enabled=True)
            return redirect('task_list')
    else:
        form = ScheduledTaskForm()
        form.fields['script'].choices = script_choices

    return render(request, 'add_task.html', {'form': form})

def DPVS(request):
    if request.method == 'GET':
        # 如果是 GET 请求，渲染 DPVS.html 页面
        return render(request, 'DPVS.html')
    elif request.method == 'POST':
        result = []
        # 去除用户输入中的空格
        vip = request.POST.get('vip', '').strip()
        vip_port = request.POST.get('vip_port', '').strip()
        real_servers = request.POST.get('real_servers', '').strip()
        real_server_port = request.POST.get('real_server_port', '').strip()
        config_filename = request.POST.get('config_filename', '').strip() 
        if not vip or not vip_port or not real_servers:
            messages.error(request, "输入不可为空")
            return render(request, 'DPVS.html')

        # 检查 VIP 地址是否在指定的网段内
        try:
            vip_address = ipaddress.ip_address(vip)           
            network = ipaddress.ip_network('172.18.15.0/24', strict=False)
            if vip_address not in network:
                messages.error(request, "VIP地址必须在172.18.15.0/24网段内")
                return render(request, 'DPVS.html')
        except ValueError:
            messages.error(request, "无效的VIP地址格式")
            return render(request, 'DPVS.html')
        real_servers_list = real_servers.split(',')
        for server in real_servers_list:
            server = server.strip()  # 去除多余的空格
            try:
                ip_address(server)
            except ValueError:
                messages.error(request, f"无效的Real Server地址格式: {server}")
                return render(request, 'DPVS.html')
        # Configuration file naming
        if not config_filename:
            config_filename = f"{vip.replace('.', '-')}-{vip_port}.conf"
        else:
            # 确保 config_filename 以 .conf 结尾
            if not config_filename.endswith('.conf'):
                config_filename += '.conf'
             
        config_path = f"/etc/keepalived/service/{config_filename}"

        if os.path.exists(config_path):
            messages.error(request, f"配置文件{config_filename}已存在,请排查!")
            return render(request, 'DPVS.html')

        # Create the configuration content
        config_content = f"""
virtual_server_group {vip}-{vip_port} {{
    {vip} {vip_port}
}}
virtual_server group {vip}-{vip_port} {{
    delay_loop 3
    lb_algo rr
    lb_kind FNAT
    protocol TCP

    laddr_group_name laddr_g1
"""

        for server in real_servers_list:
            config_content += f"""
    real_server {server.strip()} {real_server_port} {{
        weight 100
        inhibit_on_failure
        TCP_CHECK {{
            nb_sock_retry 2
            connect_timeout 3
            connect_port {real_server_port}
        }}
    }}
"""
        config_content += "}\n"
        os.chdir(global_settings.inventory_path)
        nr = InitNornir(config_file="config.yaml")
        devices = nr.filter(F(hostname="172.18.14.4") | F(hostname="172.18.14.5"))

        def configure_device(task):
            ip_output = task.run(task=netmiko_send_command, command_string="ip addr").result
            frr_config = task.run(task=netmiko_send_command, command_string="cat /etc/frr/frr.conf").result
            start_sh_content = task.run(task=netmiko_send_command, command_string="cat /root/dpvs/bin/start.sh").result

            if f"{vip}/32" in ip_output:
                messages.error(request, f"VIP {vip}/32已经存在,请再次确认!")
                return

            commands = [
                f"/root/dpvs/bin/dpip addr add {vip}/32 dev bond0",
                f"ip addr add {vip}/32 dev bond0.kni"
            ]
            task.run(task=netmiko_send_config, config_commands=commands)
            result.append(f"{task.host}: VIP已添加")
            # Write config_content to config_path line by line
            lines = config_content.splitlines()
            if lines:
                task.run(task=netmiko_send_command, command_string=f"echo '{lines[0]}' > {config_path}")
                for line in lines[1:]:
                    task.run(task=netmiko_send_command, command_string=f"echo '{line}' >> {config_path}")
            result.append(f"{task.host}: keepalived配置文件已完成")

            lines = frr_config.splitlines()
            seq_numbers = [
                int(line.split()[4]) for line in lines if line.strip().startswith("ip prefix-list ALLOWED seq") and line.split()[4].isdigit()
            ]
            max_seq = max(seq_numbers) if seq_numbers else 0
            new_seq = max_seq + 5

            new_lines = []
            last_prefix_list_index = None

            for index, line in enumerate(lines):
                new_lines.append(line)
                if line.strip() == "address-family ipv4 unicast":
                    new_lines.append(f"  network {vip}/32")
                if line.strip().startswith("ip prefix-list ALLOWED"):
                    last_prefix_list_index = index

            if last_prefix_list_index is not None:
                new_lines.insert(last_prefix_list_index + 1, f"ip prefix-list ALLOWED seq {new_seq} permit {vip}/32")
            else:
                raise Exception("未能找到任何 'ip prefix-list ALLOWED' 行。")

            new_frr_conf_content = "\n".join(new_lines)
            new_frr_conf_lines = new_frr_conf_content.splitlines()

            if new_frr_conf_lines:
                task.run(task=netmiko_send_command, command_string=f"echo '{new_frr_conf_lines[0]}' > /etc/frr/frr.conf")
                for line in new_frr_conf_lines[1:]:
                    task.run(task=netmiko_send_command, command_string=f"echo '{line}' >> /etc/frr/frr.conf")
            result.append(f"{task.host}: 完成bgp相关配置")

            start_sh_lines = start_sh_content.splitlines()
            insert_before_line = "/root/dpvs/bin/keepalived -f /etc/keepalived/keepalived.conf"
            insert_index = next((i for i, line in enumerate(start_sh_lines) if insert_before_line in line), None)

            if insert_index is not None:
                start_sh_lines.insert(insert_index, f"/root/dpvs/bin/dpip addr add {vip}/32 dev bond0")
                start_sh_lines.insert(insert_index + 1, f"ip addr add {vip}/32 dev bond0.kni")

            new_start_sh_content = "\n".join(start_sh_lines)
            start_sh_lines = new_start_sh_content.splitlines()

            if start_sh_lines:
                task.run(task=netmiko_send_command, command_string=f"echo '{start_sh_lines[0]}' > /root/dpvs/bin/start.sh")
                for line in start_sh_lines[1:]:
                    task.run(task=netmiko_send_command, command_string=f"echo '{line}' >> /root/dpvs/bin/start.sh")

            result.append(f"{task.host}: 完成开机脚本start.sh的更新")

            task.run(task=netmiko_send_command, command_string="vtysh -f /etc/frr/frr.conf")
            task.run(task=netmiko_send_command, command_string="vtysh -c 'write'")
            task.run(task=netmiko_send_command, command_string="systemctl reload keepalived")
            result.append("验证信息如下:")
            YZ1 = task.run(task=netmiko_send_command, command_string="ip addr | grep 172.18.15.").result
            result.append(YZ1)
            YZ2 = task.run(task=netmiko_send_command, command_string="vtysh -c 'show ip route | in 172.18.15.'").result
            result.append(YZ2)

        devices.run(task=configure_device)
        result = '\n'.join(result)

    return render(request, 'verify_config.html', {'result': result})


def query_vip(request):
    if request.method == 'GET':
        try:
            # 初始化 Nornir
            os.chdir(global_settings.inventory_path)
            nr = InitNornir(config_file="config.yaml")

            # 过滤出特定的设备
            device = nr.filter(hostname="172.18.14.4")

            # 定义任务
            def get_ip_addr(task: Task) -> Result:
                result = task.run(task=netmiko_send_command, command_string="ip addr")
                return Result(host=task.host, result=result.result)

            # 执行任务
            result = device.run(task=get_ip_addr)

            # 提取结果
            ip_output = list(result.values())[0][0].result

            # 提取已占用的 VIP 地址
            occupied_vips = set()
            for line in ip_output.splitlines():
                match = re.search(r'inet (172\.18\.15\.\d+)/\d+', line)
                if match:
                    occupied_vips.add(match.group(1))

            # 定义可用的 VIP 地址范围
            network = ip_network('172.18.15.0/24', strict=False)
            available_vips = []

            # 计算未被占用的 VIP 地址
            for ip in network.hosts():
                ip_str = str(ip)
                if ip_str not in occupied_vips and ip_str != '172.18.15.0' and ip_str != '172.18.15.255' and ip_str != '172.18.15.1':
                    available_vips.append(ip_str)

            # 合并连续的 IP 地址为范围
            def merge_ip_ranges(ip_list):
                if not ip_list:
                    return []

                ip_list = sorted(ip_list, key=lambda x: int(x.split('.')[-1]))
                ranges = []
                start = ip_list[0]
                end = ip_list[0]

                for ip in ip_list[1:]:
                    if int(ip.split('.')[-1]) == int(end.split('.')[-1]) + 1:
                        end = ip
                    else:
                        if start == end:
                            ranges.append(start)
                        else:
                            ranges.append(f"{start}-{end}")
                        start = ip
                        end = ip

                if start == end:
                    ranges.append(start)
                else:
                    ranges.append(f"{start}-{end}")

                return ranges

            ip_ranges = merge_ip_ranges(available_vips)
            return JsonResponse({'success': True, 'vips': ip_ranges})

        except Exception as e:
            print(f"连接到设备失败: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

def disk_replace(request):
    if request.method == 'GET':
        return render(request, 'disk_replace.html')
    elif request.method == 'POST':
        result = []
        ip = request.POST.get('ip', '').strip()
        # 初始化 Nornir
        os.chdir(global_settings.inventory_path)
        nr = InitNornir(config_file="config.yaml")

        # 过滤出特定的设备
        device = nr.filter(hostname=ip)

        # 定义任务
        def get_ip_addr(task: Task) -> Result:
            result = task.run(task=netmiko_send_command, command_string="python  megaraid_status.py")
            return Result(host=task.host, result=result.result)

        # 执行任务
        result = device.run(task=get_ip_addr)

        # 提取结果
        output = list(result.values())[0][0].result
        lines = output.splitlines()
        failed_disk_info = {}
        array_info = {}
        
        for line in lines:
            if "Failed" in line:
                parts = line.split('|')
                failed_disk_info = {
                    'slot_id': parts[7].strip(),
                    'size': parts[3].strip(),
                    'status': parts[4].strip(),
                    'id': parts[0].strip()
                }
        
        # 查找对应的Array信息
        for line in lines:
            if failed_disk_info.get('id').replace('p', '') in line and "Offline" in line:
                parts = line.split('|')
                array_info = {
                    'type': parts[1].strip(),
                    'os_path': parts[7].strip()
                }
        return render(request, 'disk_replace.html', {
            'form': form,
            'failed_disk_info': failed_disk_info,
            'array_info': array_info
        })
    return render(request, 'disk_replace.html', {'form': form})
